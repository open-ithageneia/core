"""Generate a sample Drag-and-Drop import ZIP file.

The output is a ``.zip`` archive (under ``fixtures/``) containing:

- ``drag_and_drop.xlsx`` — spreadsheet with drag-and-drop quiz data.

The spreadsheet columns are:

| id | category | left_title | right_title | left_values | right_values |

``left_values`` and ``right_values`` are **comma-separated** lists.
Each value is placed into the corresponding column bucket.

Usage:
    uv run python scripts/generate_sample_drag_and_drop_xlsx.py
"""

import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# -- column definitions -------------------------------------------------------

HEADERS = [
	"id",
	"category",
	"left_title",
	"right_title",
	"left_values",
	"right_values",
]

# -- sample rows ---------------------------------------------------------------

ROWS = [
	# Row 1: Greek geography — rivers vs lakes
	{
		"id": "",
		"category": "GEOGRAPHY",
		"left_title": "Ποταμοί",
		"right_title": "Λίμνες",
		"left_values": "Αλιάκμονας, Πηνειός, Εβρος, Αχελώος",
		"right_values": "Κερκίνη, Τριχωνίδα, Βιστωνίδα, Κορώνεια",
	},
	# Row 2: Civics — rights vs obligations
	{
		"id": "",
		"category": "CIVICS",
		"left_title": "Rights",
		"right_title": "Obligations",
		"left_values": "Vote, Free speech, Education",
		"right_values": "Pay taxes, Military service, Obey laws",
	},
	# Row 3: History — ancient vs modern
	{
		"id": "",
		"category": "HISTORY",
		"left_title": "Ancient Greece",
		"right_title": "Modern Greece",
		"left_values": "Battle of Marathon, Peloponnesian War",
		"right_values": "War of Independence, Balkan Wars",
	},
	# Row 4: Culture — food vs music
	{
		"id": "",
		"category": "CULTURE",
		"left_title": "Traditional food",
		"right_title": "Traditional music",
		"left_values": "Moussaka, Souvlaki, Spanakopita",
		"right_values": "Rebetiko, Dimotika, Nisiotika",
	},
]

# -- build workbook ------------------------------------------------------------


def build_workbook() -> Workbook:
	wb = Workbook()
	ws = wb.active
	ws.title = "Drag and Drop"

	# Write header row
	for col_idx, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = HEADER_FONT
		cell.fill = HEADER_FILL
		cell.alignment = HEADER_ALIGNMENT

	# Write data rows
	for row_idx, row_data in enumerate(ROWS, start=2):
		for col_idx, header in enumerate(HEADERS, start=1):
			value = row_data.get(header, "")
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.alignment = CELL_ALIGNMENT

	# Auto-size columns
	for col_idx, header in enumerate(HEADERS, start=1):
		col_letter = ws.cell(row=1, column=col_idx).column_letter
		max_len = len(header)
		for row_idx in range(2, len(ROWS) + 2):
			val = str(ws.cell(row=row_idx, column=col_idx).value or "")
			max_len = max(max_len, min(len(val), 60))
		ws.column_dimensions[col_letter].width = max_len + 4

	# Freeze header row
	ws.freeze_panes = "A2"

	return wb


def build_zip(wb: Workbook) -> bytes:
	"""Package the workbook into a ZIP archive."""
	buf = io.BytesIO()

	with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
		xlsx_buf = io.BytesIO()
		wb.save(xlsx_buf)
		zf.writestr("drag_and_drop.xlsx", xlsx_buf.getvalue())

	return buf.getvalue()


if __name__ == "__main__":
	wb = build_workbook()

	zip_path = "fixtures/sample_drag_and_drop.zip"
	zip_bytes = build_zip(wb)
	with open(zip_path, "wb") as f:
		f.write(zip_bytes)

	print(f"✅ Created {zip_path}")
	print(f"   Contains: drag_and_drop.xlsx ({len(ROWS)} sample rows)")

