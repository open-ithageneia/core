"""Generate a sample Fill-in-the-Blank import ZIP file.

The output is a ``.zip`` archive (under ``fixtures/``) containing:

- ``fill_in_the_blank.xlsx`` — spreadsheet with fill-in-the-blank quiz data.
- ``images/`` — optional folder with image files referenced by the
  ``prompt_image`` column.

The spreadsheet columns are:

| id | category | show_answers_as_choices | prompt_image | text_1 | text_2 | ... |

Blank syntax inside ``text_N`` columns::

    The <{{correct answer}}*> is the only choice.
    The <{{correct}}*, {{wrong1}}, {{wrong2}}> Amendment abolished slavery.

- ``<...>`` marks a blank.
- ``{{choice}}`` is an option shown to the user.
- ``*`` after the closing ``}}`` marks the correct answer (exactly one).

Usage:
    uv run python scripts/generate_sample_fill_in_the_blank_xlsx.py
"""

import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# -- placeholder image for prompt_image column --------------------------------

_IMAGE_FILES: dict[str, bytes] = {}


def _make_placeholder_image(color: str, filename: str) -> str:
	"""Create an 80×80 PNG, stash its bytes, and return the filename."""
	img = Image.new("RGB", (80, 80), color)
	buf = io.BytesIO()
	img.save(buf, format="PNG")
	_IMAGE_FILES[filename] = buf.getvalue()
	return filename


PROMPT_IMAGE = _make_placeholder_image("#8b5cf6", "prompt_purple.png")

# -- column definitions -------------------------------------------------------

HEADERS = [
	"id",
	"category",
	"show_answers_as_choices",
	"prompt_image",
	"text_1",
	"text_2",
	"text_3",
]

# -- sample rows ---------------------------------------------------------------

ROWS = [
	# Row 1: single blank, no choices shown, no image
	{
		"id": "",
		"category": "HISTORY",
		"show_answers_as_choices": "false",
		"prompt_image": "",
		"text_1": "The <{{13th}}*, {{Thirteenth}}> Amendment abolished slavery.",
	},
	# Row 2: multiple texts, show answers as choices
	{
		"id": "",
		"category": "CIVICS",
		"show_answers_as_choices": "true",
		"prompt_image": "",
		"text_1": "<{{Congress}}*> makes the laws.",
		"text_2": "The president lives in the <{{White House}}*>.",
	},
	# Row 3: multiple blanks per text, with a prompt image
	{
		"id": "",
		"category": "GEOGRAPHY",
		"show_answers_as_choices": "false",
		"prompt_image": PROMPT_IMAGE,
		"text_1": "The capital of <{{France}}*> is <{{Paris}}*>.",
		"text_2": "<{{The Amazon}}*, {{Amazon River}}> is the longest river in South America.",
		"text_3": "The <{{Pacific}}*> is the largest ocean.",
	},
	# Row 4: single blank, show as choices
	{
		"id": "",
		"category": "CULTURE",
		"show_answers_as_choices": "true",
		"prompt_image": "",
		"text_1": "<{{Shakespeare}}*> wrote Romeo and Juliet.",
	},
]

# -- build workbook ------------------------------------------------------------


def build_workbook() -> Workbook:
	wb = Workbook()
	ws = wb.active
	ws.title = "Fill in the Blank"

	# Write header row
	for col_idx, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = HEADER_FONT
		cell.fill = HEADER_FILL
		cell.alignment = HEADER_ALIGNMENT

	# Write data rows
	for row_idx, row_data in enumerate(ROWS, start=2):
		for col_idx, header in enumerate(HEADERS, start=1):
			value = row_data.get(header, "")
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.alignment = CELL_ALIGNMENT

	# Auto-size columns
	for col_idx, header in enumerate(HEADERS, start=1):
		col_letter = ws.cell(row=1, column=col_idx).column_letter
		max_len = len(header)
		for row_idx in range(2, len(ROWS) + 2):
			val = str(ws.cell(row=row_idx, column=col_idx).value or "")
			max_len = max(max_len, min(len(val), 70))
		ws.column_dimensions[col_letter].width = max_len + 4

	# Freeze header row
	ws.freeze_panes = "A2"

	return wb


def build_zip(wb: Workbook) -> bytes:
	"""Package the workbook and images into a ZIP archive."""
	buf = io.BytesIO()

	with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
		xlsx_buf = io.BytesIO()
		wb.save(xlsx_buf)
		zf.writestr("fill_in_the_blank.xlsx", xlsx_buf.getvalue())

		for filename, image_bytes in _IMAGE_FILES.items():
			zf.writestr(f"images/{filename}", image_bytes)

	return buf.getvalue()


if __name__ == "__main__":
	wb = build_workbook()

	zip_path = "fixtures/sample_fill_in_the_blank.zip"
	zip_bytes = build_zip(wb)
	with open(zip_path, "wb") as f:
		f.write(zip_bytes)

	print(f"✅ Created {zip_path}")
	print(
		f"   Contains: fill_in_the_blank.xlsx ({len(ROWS)} sample rows)"
		f" + {len(_IMAGE_FILES)} image(s) in images/"
	)

