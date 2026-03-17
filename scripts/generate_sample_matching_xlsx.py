"""Generate a sample Matching import ZIP file.

The output is a ``.zip`` archive (under ``fixtures/``) containing:

- ``matching.xlsx`` — spreadsheet with matching quiz data.

The spreadsheet columns are:

| id | category | left_title | right_title | items |

``items`` is a **comma-separated** list of pairs in ``left/right`` format.
Items are matched **by position**: the 1st pair becomes left item 1 and
right item 1, etc.

Usage:
    uv run python scripts/generate_sample_matching_xlsx.py
"""

import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# -- column definitions -------------------------------------------------------

HEADERS = [
	"id",
	"category",
	"left_title",
	"right_title",
	"items",
]

# -- sample rows ---------------------------------------------------------------

ROWS = [
	# Row 1: Greek geography regions → prefectures
	{
		"id": "",
		"category": "GEOGRAPHY",
		"left_title": "Γεωγραφικά διαμερίσματα",
		"right_title": "Περιφερειακές ενότητες",
		"items": "Στερεά Ελλάδα - Εύβοια/Βοιωτίας, Πελοπόννησος/Αρκαδίας, Kρήτη/Λασιθίου, Νησιά Αιγαίου/Λέσβου",
	},
	# Row 2: Civics — institutions → roles
	{
		"id": "",
		"category": "CIVICS",
		"left_title": "Institution",
		"right_title": "Role",
		"items": "Parliament/Legislates, Supreme Court/Interprets law, President/Head of state",
	},
	# Row 3: History — dates → events
	{
		"id": "",
		"category": "HISTORY",
		"left_title": "Date",
		"right_title": "Event",
		"items": "1821/Greek War of Independence, 1940/OXI Day, 1974/Restoration of Democracy",
	},
	# Row 4: Culture — landmarks → cities
	{
		"id": "",
		"category": "CULTURE",
		"left_title": "Landmark",
		"right_title": "City",
		"items": "Parthenon/Athens, White Tower/Thessaloniki, Palace of Knossos/Heraklion",
	},
]

# -- build workbook ------------------------------------------------------------


def build_workbook() -> Workbook:
	wb = Workbook()
	ws = wb.active
	ws.title = "Matching"

	# Write header row
	for col_idx, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = HEADER_FONT
		cell.fill = HEADER_FILL
		cell.alignment = HEADER_ALIGNMENT

	# Write data rows
	for row_idx, row_data in enumerate(ROWS, start=2):
		for col_idx, header in enumerate(HEADERS, start=1):
			value = row_data.get(header, "")
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.alignment = CELL_ALIGNMENT

	# Auto-size columns
	for col_idx, header in enumerate(HEADERS, start=1):
		col_letter = ws.cell(row=1, column=col_idx).column_letter
		max_len = len(header)
		for row_idx in range(2, len(ROWS) + 2):
			val = str(ws.cell(row=row_idx, column=col_idx).value or "")
			max_len = max(max_len, min(len(val), 60))
		ws.column_dimensions[col_letter].width = max_len + 4

	# Freeze header row
	ws.freeze_panes = "A2"

	return wb


def build_zip(wb: Workbook) -> bytes:
	"""Package the workbook into a ZIP archive."""
	buf = io.BytesIO()

	with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
		xlsx_buf = io.BytesIO()
		wb.save(xlsx_buf)
		zf.writestr("matching.xlsx", xlsx_buf.getvalue())

	return buf.getvalue()


if __name__ == "__main__":
	wb = build_workbook()

	zip_path = "fixtures/sample_matching.zip"
	zip_bytes = build_zip(wb)
	with open(zip_path, "wb") as f:
		f.write(zip_bytes)

	print(f"✅ Created {zip_path}")
	print(f"   Contains: matching.xlsx ({len(ROWS)} sample rows)")
