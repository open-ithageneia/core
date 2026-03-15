"""Generate a sample Statement import ZIP file.

The output is a ``.zip`` archive (under ``fixtures/``) containing:

- ``statements.xlsx`` — spreadsheet with quiz data.  Image columns hold
  **filenames** (e.g. ``prompt_blue.png``) that reference files in the
  ``images/`` folder.
- ``images/`` — folder with the actual image files.

Usage:
    uv run python scripts/generate_sample_statements_xlsx.py
"""

import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# -- tiny placeholder images --------------------------------------------------

_IMAGE_FILES: dict[str, bytes] = {}


def _make_placeholder_image(color: str, filename: str) -> str:
	"""Create an 80×80 PNG, stash its bytes, and return the filename."""
	img = Image.new("RGB", (80, 80), color)
	buf = io.BytesIO()
	img.save(buf, format="PNG")
	_IMAGE_FILES[filename] = buf.getvalue()
	return filename


PROMPT_IMAGE = _make_placeholder_image("#3b82f6", "prompt_blue.png")
CHOICE_IMAGE_A = _make_placeholder_image("#22c55e", "choice_green.png")
CHOICE_IMAGE_B = _make_placeholder_image("#ef4444", "choice_red.png")

# -- column definitions -------------------------------------------------------

HEADERS = [
	"id",
	"type",
	"category",
	"prompt_text",
	"prompt_image",
	"choice1_text",
	"choice1_image",
	"choice1_is_correct",
	"choice2_text",
	"choice2_image",
	"choice2_is_correct",
	"choice3_text",
	"choice3_image",
	"choice3_is_correct",
	"choice4_text",
	"choice4_image",
	"choice4_is_correct",
]

# -- sample rows ---------------------------------------------------------------

ROWS = [
	# Row 1: TRUE_FALSE — text only, no images
	{
		"id": "",
		"type": "TRUE_FALSE",
		"category": "GEOGRAPHY",
		"prompt_text": "The capital of Greece is Athens.",
		"prompt_image": "",
		"choice1_text": "True",
		"choice1_image": "",
		"choice1_is_correct": "true",
		"choice2_text": "False",
		"choice2_image": "",
		"choice2_is_correct": "false",
	},
	# Row 2: TRUE_FALSE — with a prompt image
	{
		"id": "",
		"type": "TRUE_FALSE",
		"category": "CIVICS",
		"prompt_text": "This flag belongs to Italy.",
		"prompt_image": PROMPT_IMAGE,
		"choice1_text": "True",
		"choice1_image": "",
		"choice1_is_correct": "false",
		"choice2_text": "False",
		"choice2_image": "",
		"choice2_is_correct": "true",
	},
	# Row 3: MULTIPLE_CHOICE — text only
	{
		"id": "",
		"type": "MULTIPLE_CHOICE",
		"category": "HISTORY",
		"prompt_text": "Who was the first President of the United States?",
		"prompt_image": "",
		"choice1_text": "Thomas Jefferson",
		"choice1_image": "",
		"choice1_is_correct": "false",
		"choice2_text": "George Washington",
		"choice2_image": "",
		"choice2_is_correct": "true",
		"choice3_text": "Abraham Lincoln",
		"choice3_image": "",
		"choice3_is_correct": "false",
		"choice4_text": "John Adams",
		"choice4_image": "",
		"choice4_is_correct": "false",
	},
	# Row 4: MULTIPLE_CHOICE — choices with images
	{
		"id": "",
		"type": "MULTIPLE_CHOICE",
		"category": "CULTURE",
		"prompt_text": "Which image shows the Parthenon?",
		"prompt_image": "",
		"choice1_text": "Option A",
		"choice1_image": CHOICE_IMAGE_A,
		"choice1_is_correct": "true",
		"choice2_text": "Option B",
		"choice2_image": CHOICE_IMAGE_B,
		"choice2_is_correct": "false",
	},
	# Row 5: MULTIPLE_CHOICE — prompt image + choice images
	{
		"id": "",
		"type": "MULTIPLE_CHOICE",
		"category": "GEOGRAPHY",
		"prompt_text": "Identify the landmark shown in the image.",
		"prompt_image": PROMPT_IMAGE,
		"choice1_text": "Eiffel Tower",
		"choice1_image": CHOICE_IMAGE_A,
		"choice1_is_correct": "false",
		"choice2_text": "Acropolis",
		"choice2_image": CHOICE_IMAGE_B,
		"choice2_is_correct": "true",
		"choice3_text": "Colosseum",
		"choice3_image": "",
		"choice3_is_correct": "false",
	},
]

# -- build workbook ------------------------------------------------------------


def build_workbook() -> Workbook:
	wb = Workbook()
	ws = wb.active
	ws.title = "Statements"

	# Write header row
	for col_idx, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = HEADER_FONT
		cell.fill = HEADER_FILL
		cell.alignment = HEADER_ALIGNMENT

	# Write data rows
	for row_idx, row_data in enumerate(ROWS, start=2):
		for col_idx, header in enumerate(HEADERS, start=1):
			value = row_data.get(header, "")
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.alignment = CELL_ALIGNMENT

	# Auto-size text columns, give image columns a reasonable width
	IMAGE_COLS = {h for h in HEADERS if "image" in h}
	for col_idx, header in enumerate(HEADERS, start=1):
		col_letter = ws.cell(row=1, column=col_idx).column_letter
		if header in IMAGE_COLS:
			ws.column_dimensions[col_letter].width = 22
		else:
			max_len = len(header)
			for row_idx in range(2, len(ROWS) + 2):
				val = str(ws.cell(row=row_idx, column=col_idx).value or "")
				max_len = max(max_len, min(len(val), 50))
			ws.column_dimensions[col_letter].width = max_len + 4

	# Freeze header row
	ws.freeze_panes = "A2"

	return wb


def build_zip(wb: Workbook) -> bytes:
	"""Package the workbook and placeholder images into a ZIP archive."""
	buf = io.BytesIO()

	with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
		# Write the spreadsheet
		xlsx_buf = io.BytesIO()
		wb.save(xlsx_buf)
		zf.writestr("statements.xlsx", xlsx_buf.getvalue())

		# Write all referenced images into images/ folder
		for filename, image_bytes in _IMAGE_FILES.items():
			zf.writestr(f"images/{filename}", image_bytes)

	return buf.getvalue()


if __name__ == "__main__":
	wb = build_workbook()

	zip_path = "fixtures/sample_statements.zip"
	zip_bytes = build_zip(wb)
	with open(zip_path, "wb") as f:
		f.write(zip_bytes)

	print(f"✅ Created {zip_path}")
	print(f"   Contains: statements.xlsx + {len(_IMAGE_FILES)} image(s) in images/")
